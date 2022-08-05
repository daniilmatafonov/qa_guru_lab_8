from openpyxl import load_workbook


def test_read_xlsx():
    workbook = load_workbook('unpacked/names.xlsx')
    sheet = workbook.active
    name = sheet.cell(row=4, column=1).value
    assert 'Jane' == name